from decimal import Decimal

from dominate.tags import header as header_tag, table, tbody, td, th, thead, tr
from dominate.util import raw
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from trytond.modules.html_report.dominate_report import DominateReport
from trytond.transaction import Transaction
from trytond.modules.xgettext import _
from trytond.modules.html_report.tools import save_virtual_workbook

from ..report import Report


class FinancialStatementBase(DominateReport):
    _landscape_threshold = 5
    side_margin = 0.5

    @classmethod
    def language(cls, records):
        return Transaction().language or 'en'

    @classmethod
    def css(cls, action, data, records):
        css = super().css(action, data, records) or ''
        if records:
            record, = records
            if len(cls._comparison_periods(record)) > cls._landscape_threshold:
                css += '\n@page { size: A4 landscape; }\n'
        return css

    @staticmethod
    def _raw(record):
        return getattr(record, 'raw', record)

    @classmethod
    def _record_id(cls, record):
        return cls._raw(record).id

    @classmethod
    def _display_value(cls, record, field, fallback=None):
        raw_record = cls._raw(record)
        rendered = getattr(record, 'render', None)
        if rendered is not None and hasattr(rendered, field):
            return getattr(rendered, field)
        if hasattr(raw_record, field):
            return getattr(raw_record, field)
        if fallback and hasattr(raw_record, fallback):
            return getattr(raw_record, fallback)
        return ''

    @classmethod
    def _comparison_periods(cls, record):
        return Report._ordered_periods(cls._raw(record))

    @classmethod
    def _period_label(cls, period):
        return cls._display_value(period.fiscalyear, 'name', fallback='rec_name')

    @classmethod
    def _period_header_label(cls, index, period):
        return '%s %s' % (
            _('Period'),
            index,
            )

    @classmethod
    def _period_range(cls, period):
        period_raw = cls._raw(period)
        if period_raw.start_period and period_raw.end_period:
            return '%s - %s' % (
                cls._display_value(period.start_period, 'rec_name', fallback='name'),
                cls._display_value(period.end_period, 'rec_name', fallback='name'),
                )
        return '%s (%s)' % (
            cls._display_value(period.fiscalyear, 'rec_name', fallback='name'),
            _('Full fiscal year'),
            )

    @classmethod
    def _format_amount(cls, value, currency):
        if value in (None, ''):
            return ''
        return cls.format_currency(value, None, currency, symbol=False)

    @classmethod
    def _line_key(cls, line):
        raw_line = cls._raw(line)
        if raw_line.template_line:
            return ('template', cls._record_id(raw_line.template_line))
        return ('code', raw_line.code)

    @classmethod
    def _period_line_lookup(cls, period):
        return {cls._line_key(line): line for line in period.lines}

    @classmethod
    def _structural_lines(cls, record):
        periods = cls._comparison_periods(record)
        if not periods:
            return []
        return list(periods[0].lines)

    @classmethod
    def _line_column_value(cls, line, column):
        period_line = column['lookup'].get(cls._line_key(line))
        if not period_line:
            return ''
        currency = cls._raw(period_line.currency) if period_line.currency else None
        return cls._format_amount(cls._raw(period_line).value, currency)

    @classmethod
    def _detail_rows(cls, line, record, columns):
        by_account = {}
        for column in columns:
            period_line = column['lookup'].get(cls._line_key(line))
            if not period_line:
                continue
            for value in period_line.line_accounts:
                amount = cls._raw(value).balance or Decimal(0)
                key = cls._record_id(value.account)
                entry = by_account.setdefault(key, {
                        'label': '%s - %s' % (
                            getattr(value.account, 'render', value.account).code,
                            getattr(value.account, 'render', value.account).name,
                            ),
                        'values': {},
                        })
                entry['values'][column['id']] = amount
        rows = []
        for entry in sorted(by_account.values(),
                key=lambda item: item['label']):
            if not any(entry['values'].values()):
                continue
            formatted = {}
            for column in columns:
                amount = entry['values'].get(column['id'], Decimal(0))
                if amount is None:
                    continue
                formatted[column['id']] = cls._format_amount(
                    amount, cls._raw(line.currency) if line.currency else None)
            if formatted:
                entry['values'] = formatted
                rows.append(entry)
        return rows

    @classmethod
    def header(cls, action, data, records):
        record, = records
        header_node = header_tag(id='header')
        with header_node:
            with table(style='width:100%;'):
                with tr():
                    with td():
                        raw('<h2><i>%s</i></h2>' % record.company.render.rec_name)
                        raw('<h3>NIF: %s</h3>' % (
                            record.company.party.tax_identifier
                            and record.company.party.tax_identifier.render.code
                            or '-'))
                    with td():
                        raw('<h1>%s</h1>' % record.render.rec_name)
        return header_node

    @classmethod
    def title(cls, action, data, records):
        record, = records
        return record.render.rec_name

    @classmethod
    def _table_columns(cls, record):
        return [{
                'id': cls._record_id(period),
                'label': cls._period_range(period),
                'period': period,
                'lookup': cls._period_line_lookup(period),
                }
            for period in cls._comparison_periods(record)]

    @classmethod
    def _build_table(cls, record, include_details):
        columns = cls._table_columns(record)
        with table(style='page-break-inside: auto;') as table_node:
            with thead(style='border-bottom: 1px solid black; border-top: 1px solid black'):
                with tr(style='page-break-inside:avoid;'):
                    th(_('Concept'), nowrap=True)
                    for column in columns:
                        th(column['label'], nowrap=True,
                            style='text-align: right')
            with tbody():
                for line in cls._structural_lines(record):
                    raw_line = cls._raw(line)
                    if not raw_line.visible:
                        continue
                    line_style = 'page-break-inside:avoid;'
                    if raw_line.page_break:
                        line_style += ' page-break-after:always;'
                    with tr(style=line_style):
                        if not raw_line.parent:
                            th(cls._display_value(line, 'name', fallback='rec_name'))
                        else:
                            td(cls._display_value(line, 'name', fallback='rec_name'))
                        for column in columns:
                            td(cls._line_column_value(line, column),
                                style='text-align: right')
                    if include_details:
                        for detail in cls._detail_rows(line, record, columns):
                            detail_style = 'page-break-inside:avoid; color: #A2A2A2;'
                            if raw_line.page_break:
                                detail_style += ' page-break-after:always;'
                            with tr(style=detail_style):
                                td(detail['label'])
                                for column in columns:
                                    td(detail['values'].get(column['id'], ''),
                                        style='text-align: right')
        return table_node

    @classmethod
    def _summary_lines(cls, record):
        return [line for line in cls._structural_lines(record)
            if cls._raw(line).visible]

    @classmethod
    def _line_numeric_value(cls, line, column):
        period_line = column['lookup'].get(cls._line_key(line))
        if not period_line:
            return None
        return cls._raw(period_line).value


class FinancialStatementExport(FinancialStatementBase):
    __name__ = 'account.financial.statement.export'
    _number_format = '#,##0.00'

    @classmethod
    def execute(cls, ids, data):
        if not ids:
            return
        action, model = cls.get_action(data)
        cls.check_access(action, model, ids)
        records = cls._get_records(ids, model, data)
        workbook = Workbook()
        workbook.remove(workbook.active)
        for record in records:
            cls._add_sheet(workbook, record)
        name = records[0].rec_name if len(records) == 1 else action.name
        return ('xlsx', save_virtual_workbook(workbook), False, name)

    @classmethod
    def _sheet_title(cls, record):
        name = cls._display_value(record, 'name', fallback='rec_name') or 'Report'
        return name.replace('/', '_').replace(':', '_')[:31]

    @classmethod
    def _add_sheet(cls, workbook, record):
        sheet = workbook.create_sheet(cls._sheet_title(record))
        columns = cls._table_columns(record)
        headers = [_('Name')] + [
            cls._period_range(column['period'])
            or column['label']
            for column in columns]
        title = cls._display_value(record, 'name', fallback='rec_name') or 'Report'
        border_side = Side(style='thin')

        sheet.append([title])
        if len(headers) > 1:
            sheet.merge_cells(start_row=1, start_column=1,
                end_row=1, end_column=len(headers))
        for column_index in range(1, len(headers) + 1):
            cell = sheet.cell(1, column_index)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                top=border_side,
                bottom=border_side,
                left=border_side if column_index == 1 else None,
                right=border_side if column_index == len(headers) else None,
                )

        sheet.append(headers)
        for cell in sheet[2]:
            cell.font = Font(bold=True)
            cell.border = Border(
                top=border_side,
                bottom=border_side,
                left=border_side,
                right=border_side,
                )

        for line in cls._summary_lines(record):
            row = [cls._display_value(line, 'name', fallback='rec_name')]
            row.extend(cls._line_numeric_value(line, column) for column in columns)
            sheet.append(row)
            row_index = sheet.max_row
            for column_index in range(2, len(headers) + 1):
                cell = sheet.cell(row_index, column_index)
                if cell.value is not None:
                    cell.number_format = cls._number_format

        widths = {1: 42}
        for index in range(2, len(headers) + 1):
            widths[index] = 18
        for index, width in widths.items():
            sheet.column_dimensions[get_column_letter(index)].width = width


class FinancialStatementReport(FinancialStatementBase):
    'Financial Statement Report'
    __name__ = 'account.financial.statement.report'

    @classmethod
    def body(cls, action, data, records):
        record, = records
        return cls._build_table(record, include_details=False)


class FinancialStatementDetailReport(FinancialStatementBase):
    'Financial Statement Detail Report'
    __name__ = 'account.financial.statement.detail.report'

    @classmethod
    def body(cls, action, data, records):
        record, = records
        return cls._build_table(record, include_details=True)
