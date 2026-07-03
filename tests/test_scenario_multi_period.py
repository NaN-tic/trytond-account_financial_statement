import unittest
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from proteus import Model
from trytond.exceptions import UserError
from trytond.modules.account.tests.tools import (
    create_chart,
    create_fiscalyear,
    get_accounts,
)
from trytond.modules.company.tests.tools import create_company, get_company
from trytond.pool import Pool
from trytond.tests.test_tryton import drop_db
from trytond.tests.tools import activate_modules
from trytond.transaction import Transaction


class TestFinancialStatementMultiPeriod(unittest.TestCase):

    def setUp(self):
        drop_db()
        super().setUp()

    def tearDown(self):
        drop_db()
        super().tearDown()

    def create_move(self, period, amount):
        Move = Model.get('account.move', config=self.config)
        move = Move()
        move.journal = self.journal_revenue
        move.period = period
        move.date = period.start_date
        line = move.lines.new()
        line.account = self.revenue
        line.credit = amount
        line = move.lines.new()
        line.account = self.receivable
        line.party = self.party
        line.debit = amount
        move.save()
        move.click('post')

    def setup_data(self):
        self.config = activate_modules('account_financial_statement')
        _ = create_company(config=self.config)
        self.company = get_company(config=self.config)
        create_chart(company=self.company, config=self.config)

        Party = Model.get('party.party', config=self.config)
        Journal = Model.get('account.journal', config=self.config)
        Period = Model.get('account.period', config=self.config)
        Template = Model.get('account.financial.statement.template',
            config=self.config)
        TemplateLine = Model.get('account.financial.statement.template.line',
            config=self.config)

        accounts = get_accounts(company=self.company, config=self.config)
        self.revenue = accounts['revenue']
        self.receivable = accounts['receivable']
        self.journal_revenue, = Journal.find([('code', '=', 'REV')], limit=1)
        self.party = Party(name='Customer')
        self.party.save()

        self.fiscalyears = []
        for year in (2022, 2023, 2024):
            fiscalyear = create_fiscalyear(
                company=self.company,
                today=(date(year, 1, 1), date(year, 12, 31)),
                config=self.config)
            fiscalyear.save()
            fiscalyear.click('create_period')
            self.fiscalyears.append(fiscalyear)

        for fiscalyear, amount in zip(
                self.fiscalyears,
                [Decimal('100'), Decimal('200'), Decimal('300')]):
            self.create_move(fiscalyear.periods[0], amount)

        self.adjustment_period = Period()
        self.adjustment_period.name = 'Closing'
        self.adjustment_period.start_date = self.fiscalyears[0].end_date
        self.adjustment_period.end_date = self.fiscalyears[0].end_date
        self.adjustment_period.fiscalyear = self.fiscalyears[0]
        self.adjustment_period.type = 'adjustment'
        self.adjustment_period.save()
        self.create_move(self.adjustment_period, Decimal('999'))

        self.template = Template()
        self.template.name = 'Multi Period Template'
        self.template.mode = 'credit-debit'
        line = self.template.lines.new()
        line.code = '0'
        line.name = 'Results'
        line = self.template.lines.new()
        line.code = '1'
        line.name = 'Fixed'
        line.current_value = '12.00'
        line = self.template.lines.new()
        line.code = '2'
        line.name = 'Total'
        line.current_value = 'concept("0","1")'
        self.template.save()

        revenue_line = TemplateLine()
        revenue_line.template = self.template
        revenue_line.parent = self.template.lines[0]
        revenue_line.code = 'R'
        revenue_line.name = 'Revenue'
        revenue_line.current_value = 'balance("%s")' % self.revenue.code
        revenue_line.save()

    def test(self):
        self.setup_data()
        Report = Model.get('account.financial.statement.report',
            config=self.config)

        report = Report()
        report.name = 'Three Year Report'
        report.template = self.template
        first_fiscalyear_standard_periods = [
            period for period in self.fiscalyears[0].periods
            if period.type == 'standard']
        for sequence, fiscalyear in [
                (2, self.fiscalyears[2]),
                (0, self.fiscalyears[0]),
                (1, self.fiscalyears[1])]:
            period = report.comparison_periods.new()
            period.sequence = sequence
            period.fiscalyear = fiscalyear
            if fiscalyear == self.fiscalyears[0]:
                period.start_period = first_fiscalyear_standard_periods[0]
                period.end_period = first_fiscalyear_standard_periods[-1]
        report.save()
        report.click('calculate')

        ranged_standard_report = Report()
        ranged_standard_report.name = 'Ranged Standard Report'
        ranged_standard_report.template = self.template
        ranged_standard_period = ranged_standard_report.comparison_periods.new()
        ranged_standard_period.fiscalyear = self.fiscalyears[0]
        standard_periods = [
            period for period in self.fiscalyears[0].periods
            if period.type == 'standard']
        ranged_standard_period.start_period = standard_periods[0]
        ranged_standard_period.end_period = standard_periods[-1]
        ranged_standard_report.save()
        ranged_standard_report.click('calculate')

        ranged_adjustment_report = Report()
        ranged_adjustment_report.name = 'Ranged Adjustment Report'
        ranged_adjustment_report.template = self.template
        ranged_adjustment_period = ranged_adjustment_report.comparison_periods.new()
        ranged_adjustment_period.fiscalyear = self.fiscalyears[0]
        ranged_adjustment_period.start_period = standard_periods[0]
        ranged_adjustment_period.end_period = self.adjustment_period
        ranged_adjustment_report.save()
        ranged_adjustment_report.click('calculate')

        extra_fiscalyears = []
        for year in (2021, 2025, 2026):
            fiscalyear = create_fiscalyear(
                company=self.company,
                today=(date(year, 1, 1), date(year, 12, 31)),
                config=self.config)
            fiscalyear.save()
            fiscalyear.click('create_period')
            extra_fiscalyears.append(fiscalyear)

        summary_report = Report()
        summary_report.name = 'Six Year Report'
        summary_report.template = self.template
        for fiscalyear in [
                self.fiscalyears[2], self.fiscalyears[0], extra_fiscalyears[2],
                extra_fiscalyears[0], extra_fiscalyears[1], self.fiscalyears[1]]:
            period = summary_report.comparison_periods.new()
            period.fiscalyear = fiscalyear
        summary_report.save()

        overflowing_report = Report()
        overflowing_report.name = 'Eleven Year Report'
        overflowing_report.template = self.template
        for year in range(2021, 2032):
            fiscalyear = next(
                (fiscalyear for fiscalyear in self.fiscalyears + extra_fiscalyears
                    if fiscalyear.start_date.year == year),
                None)
            if fiscalyear is None:
                fiscalyear = create_fiscalyear(
                    company=self.company,
                    today=(date(year, 1, 1), date(year, 12, 31)),
                    config=self.config)
                fiscalyear.save()
                fiscalyear.click('create_period')
                extra_fiscalyears.append(fiscalyear)
            period = overflowing_report.comparison_periods.new()
            period.fiscalyear = fiscalyear
        with self.assertRaises(UserError):
            overflowing_report.save()

        pool = Pool(self.config.database_name)
        ReportModel = pool.get('account.financial.statement.report')
        ReportLineModel = pool.get('account.financial.statement.report.line.period')
        FinancialStatementExport = pool.get(
            'account.financial.statement.export', type='report')
        FinancialStatementReport = pool.get(
            'account.financial.statement.report', type='report')
        FinancialStatementDetailReport = pool.get(
            'account.financial.statement.detail.report', type='report')

        with Transaction().start(
                self.config.database_name, self.config.user,
                context=self.config.context):
            report_record = ReportModel(report.id)
            ranged_standard_report_record = ReportModel(ranged_standard_report.id)
            ranged_adjustment_report_record = ReportModel(
                ranged_adjustment_report.id)
            summary_report_record = ReportModel(summary_report.id)

            report_period_values = {}
            for period in report_record.comparison_periods:
                revenue_line, = [line for line in period.lines if line.code == 'R']
                fixed_line, = [line for line in period.lines if line.code == '1']
                total_line, = [line for line in period.lines if line.code == '2']
                report_period_values[period.fiscalyear.id] = revenue_line.value
                self.assertEqual(fixed_line.value, Decimal('12.00'))
                self.assertEqual(total_line.value, revenue_line.value + Decimal('12.00'))
                self.assertTrue(all(
                        detail.debit is not None and detail.credit is not None
                        for detail in revenue_line.line_accounts))
            self.assertEqual(report_period_values[self.fiscalyears[0].id],
                Decimal('100.00'))
            self.assertEqual(report_period_values[self.fiscalyears[1].id],
                Decimal('200.00'))
            self.assertEqual(report_period_values[self.fiscalyears[2].id],
                Decimal('300.00'))

            ranged_standard_period_record, = (
                ranged_standard_report_record.comparison_periods)
            ranged_standard_revenue_line, = [
                line for line in ranged_standard_period_record.lines
                if line.code == 'R']
            self.assertEqual(ranged_standard_revenue_line.value, Decimal('100.00'))

            ranged_adjustment_period_record, = (
                ranged_adjustment_report_record.comparison_periods)
            ranged_adjustment_revenue_line, = [
                line for line in ranged_adjustment_period_record.lines
                if line.code == 'R']
            self.assertEqual(ranged_adjustment_revenue_line.value, Decimal('1099.00'))

            self.assertEqual(
                report_record.comparison_fiscalyears, '2022 / 2023 / 2024')
            self.assertEqual(
                summary_report_record.comparison_fiscalyears, '2021 ... 2026')
            self.assertEqual(
                len(ReportLineModel.search([
                    ('report_period.report', '=', report_record.id),
                    ('parent', '=', None),
                    ])),
                sum(len([line for line in period.lines if not line.parent])
                    for period in report_record.comparison_periods))

            oext, content, _, filename = FinancialStatementExport.execute(
                [report_record.id], {'model': 'account.financial.statement.report'})
            self.assertEqual(oext, 'xlsx')
            self.assertTrue(content.startswith(b'PK'))
            self.assertEqual(filename, report_record.rec_name)
            workbook = load_workbook(BytesIO(content))
            sheet = workbook.active
            self.assertEqual(sheet.title, report_record.name)
            exported_headers = ['Name'] + [
                FinancialStatementExport._period_range(period)
                for period in report_record.comparison_periods]
            self.assertEqual(
                [sheet.cell(2, column).value for column in range(1, 5)],
                exported_headers)
            rows = {
                sheet.cell(row, 1).value: [sheet.cell(row, column).value
                    for column in range(2, 5)]
                for row in range(3, sheet.max_row + 1)
                }
            self.assertEqual(
                rows['Revenue'],
                [Decimal('100.00'), Decimal('200.00'), Decimal('300.00')])
            self.assertEqual(sheet.cell(4, 2).number_format, '#,##0.00')
            body = str(FinancialStatementReport.body(
                    None, None, [report_record]))
            expected_header = FinancialStatementReport._period_range(
                report_record.comparison_periods[1])
            self.assertIn(
                '<th nowrap="nowrap" style="text-align: right">%s</th>'
                % expected_header,
                body)
            detail_body = str(FinancialStatementDetailReport.body(
                    None, None, [report_record]))
            self.assertIn(
                '<th nowrap="nowrap" style="text-align: right">%s</th>'
                % expected_header,
                detail_body)
            self.assertNotIn(
                '@page { size: A4 landscape; }',
                FinancialStatementReport.css(None, None, [report_record]))
            self.assertIn(
                '@page { size: A4 landscape; }',
                FinancialStatementReport.css(None, None, [summary_report_record]))
